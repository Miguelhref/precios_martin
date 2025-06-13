from flask import Flask, render_template, request, send_file
import re
from forex_python.converter import CurrencyRates
import openpyxl
import io

app = Flask(__name__)
currency = CurrencyRates()

resultados_exportables = []  # para guardar datos exportables
excel_tempfile = None  # será BytesIO (archivo en memoria)


def parse_line(line, usd_to_eur):
    match = re.search(r'(.*?)\s*\$(\d+(?:\.\d+)?)(/T)?(?:\((.*?)\))?', line)
    if match:
        modelo_raw = match.group(1).strip()
        precio_base = float(match.group(2))
        es_precio_por_t = match.group(3) == "/T"
        unidad_extra = match.group(4)

        modelo = re.sub(r'\s+', '-', modelo_raw)
        if unidad_extra:
            modelo += f"({unidad_extra})"

        cantidades = []
        opciones_match = re.search(r'(?:\s|[wW])(\d+(?:T)?(?:/\d+(?:T)?)+)', modelo_raw)
        if opciones_match:
            opciones_str = opciones_match.group(1)
            cantidades = [int(t.replace('T', '')) for t in opciones_str.split('/') if t.strip() != '']
        else:
            fallback_match = re.findall(r'(?:\s|[wW])(\d+)T?', modelo_raw)
            cantidades = [int(x) for x in fallback_match if x.isdigit()]

        precios_unitarios_raw = []
        precios_agrupados = re.search(r'\$(\d+(?:\.\d+)?(?:/\d+(?:\.\d+)?)*)[/]?T', line)
        if precios_agrupados:
            precios_str = precios_agrupados.group(1)
            precios_unitarios_raw = precios_str.split('/')
        else:
            precios_unitarios_raw = re.findall(r'\$(\d+(?:\.\d+)?)(?:/T)?', line)

        precios_unitarios = [float(p) for p in precios_unitarios_raw]

        if not cantidades and not es_precio_por_t:
            precio_usd = precio_base
            pvp_usd = round(precio_usd * 1.25, 2)
            precio_eur = round(precio_usd * usd_to_eur, 2)
            pvp_eur = round(pvp_usd * usd_to_eur, 2)
            return {
                "modelo": modelo,
                "is_multiple": False,
                "precio": f"${precio_usd}",
                "pvp": f"${pvp_usd}",
                "precio_eur": f"€{precio_eur}",
                "pvp_eur": f"€{pvp_eur}",
            }

        if not es_precio_por_t and precio_base < 30:
            es_precio_por_t = True

        if es_precio_por_t and cantidades:
            if len(precios_unitarios) < len(cantidades):
                ultimo_precio = precios_unitarios[-1] if precios_unitarios else precio_base
                while len(precios_unitarios) < len(cantidades):
                    precios_unitarios.append(ultimo_precio)
            elif len(precios_unitarios) > len(cantidades):
                precios_unitarios = precios_unitarios[:len(cantidades)]

            precios_usd = [round(cantidades[i] * precios_unitarios[i], 2) for i in range(len(cantidades))]
            precios_eur = [round(p * usd_to_eur, 2) for p in precios_usd]
            pvp_usd = [round(p * 1.25, 2) for p in precios_usd]
            pvp_eur = [round(p * usd_to_eur, 2) for p in pvp_usd]

            if len(cantidades) > 1:
                return {
                    "modelo": modelo,
                    "is_multiple": True,
                    "precio_list": [f"${precios_usd[i]} ({precios_unitarios[i]}/T)" for i in range(len(cantidades))],
                    "pvp_list": [f"${p}" for p in pvp_usd],
                    "precio_eur_list": [f"€{p}" for p in precios_eur],
                    "pvp_eur_list": [f"€{p}" for p in pvp_eur],
                }
            else:
                return {
                    "modelo": modelo,
                    "is_multiple": False,
                    "precio": f"${precios_usd[0]} ({precios_unitarios[0]}/T)",
                    "pvp": f"${pvp_usd[0]}",
                    "precio_eur": f"€{precios_eur[0]}",
                    "pvp_eur": f"€{pvp_eur[0]}",
                }

        else:
            precio_usd = precio_base
            pvp_usd = round(precio_usd * 1.25, 2)
            precio_eur = round(precio_usd * usd_to_eur, 2)
            pvp_eur = round(pvp_usd * usd_to_eur, 2)
            return {
                "modelo": modelo,
                "is_multiple": False,
                "precio": f"${precio_usd}",
                "pvp": f"${pvp_usd}",
                "precio_eur": f"€{precio_eur}",
                "pvp_eur": f"€{pvp_eur}",
            }
    return None


@app.route("/", methods=["GET", "POST"])
def index():
    global resultados_exportables, excel_tempfile
    resultados = []
    error = None
    nombre_archivo = None

    if request.method == "POST":
        try:
            usd_to_eur = currency.get_rate('USD', 'EUR')
        except:
            usd_to_eur = 0.92  # fallback

        file = request.files.get("file")
        if file:
            filename = file.filename
            nombre_archivo = filename
            filename = filename.lower()

            if filename.endswith(".txt"):
                lines = file.read().decode("utf-8").splitlines()

            elif filename.endswith(".xlsx"):
                try:
                    wb = openpyxl.load_workbook(filename=io.BytesIO(file.read()), data_only=True)
                    sheet = wb.active

                    header_row = None
                    headers = []
                    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        headers = [str(cell).strip() if cell else "" for cell in row]
                        required = ["Brand", "Model", "Hashrate/T", "Price/T", "Unit Price"]
                        if all(h in headers for h in required):
                            header_row = i
                            break
                    if not header_row:
                        error = "No se encontraron encabezados válidos en el archivo Excel."
                    else:
                        idx_brand = headers.index("Brand")
                        idx_model = headers.index("Model")
                        idx_hashrate = headers.index("Hashrate/T")
                        idx_price_per_t = headers.index("Price/T")
                        idx_unit_price = headers.index("Unit Price")

                        lines = []
                        for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
                            brand = str(row[idx_brand] or "").strip()
                            model = str(row[idx_model] or "").strip()
                            hashrate = str(row[idx_hashrate] or "").strip()
                            price_per_t = row[idx_price_per_t]
                            unit_price = row[idx_unit_price]

                            if unit_price not in [None, "", 0]:
                                line = f"{brand} {model} {hashrate} ${unit_price}"
                            elif price_per_t not in [None, "", 0]:
                                line = f"{brand} {model} {hashrate} ${price_per_t}/T"
                            else:
                                continue

                            lines.append(line)
                except Exception as e:
                    error = f"Error leyendo archivo Excel: {e}"
                    lines = []

            else:
                error = "Solo se aceptan archivos .txt o .xlsx"
                lines = []

            if not error:
                resultados_exportables = []
                for line in lines:
                    if "$" in line:
                        parsed = parse_line(line, usd_to_eur)
                        if parsed:
                            resultados.append(parsed)
                            if parsed["is_multiple"]:
                                for i in range(len(parsed["precio_list"])):
                                    resultados_exportables.append({
                                        "modelo": parsed["modelo"],
                                        "precio": parsed["precio_list"][i],
                                        "pvp": parsed["pvp_list"][i],
                                        "precio_eur": parsed["precio_eur_list"][i],
                                        "pvp_eur": parsed["pvp_eur_list"][i],
                                    })
                            else:
                                resultados_exportables.append({
                                    "modelo": parsed["modelo"],
                                    "precio": parsed["precio"],
                                    "pvp": parsed["pvp"],
                                    "precio_eur": parsed["precio_eur"],
                                    "pvp_eur": parsed["pvp_eur"],
                                })

                # Guardar Excel en memoria con BytesIO
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Modelo", "Precio (USD)", "PVP (USD)", "Precio (€)", "PVP (€)"])
                for r in resultados_exportables:
                    ws.append([r["modelo"], r["precio"], r["pvp"], r["precio_eur"], r["pvp_eur"]])

                excel_tempfile = io.BytesIO()
                wb.save(excel_tempfile)
                excel_tempfile.seek(0)

    return render_template("index.html", resultados=resultados, error=error, nombre_archivo=nombre_archivo)


@app.route("/download")
def download_excel():
    global excel_tempfile
    if excel_tempfile:
        excel_tempfile.seek(0)
        return send_file(
            excel_tempfile,
            as_attachment=True,
            download_name="resultados.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    return "No hay archivo disponible para descargar."


if __name__ == "__main__":
    app.run(debug=True)
